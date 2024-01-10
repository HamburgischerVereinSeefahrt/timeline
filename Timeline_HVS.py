import drawsvg as draw
from openpyxl import Workbook
import openpyxl

excel_file = openpyxl.load_workbook('Mappe1.xlsx')
ws = excel_file['Tabelle1']
max_rows = ws.max_row

schiffsnamen = []
anschaffungsjahre = []
verkaufsjahre = []

for i in range(2, max_rows+1):
    schiffsnamen.append(ws.cell(row = i, column = 1).value)
    anschaffungsjahre.append(ws.cell(row = i, column = 2).value)
    verkaufsjahre.append(ws.cell(row = i, column = 3).value)

print(schiffsnamen)
print(anschaffungsjahre)
print(verkaufsjahre)

d = draw.Drawing(1300, 350, origin ='top-left')
yPos = 20
for schiffsname in schiffsnamen:
    d.append(draw.Text(schiffsname, 14, 100, yPos, fill='black', font_family='sans-serif',text_anchor='end' ))
    yPos += 15




allYears = anschaffungsjahre + verkaufsjahre[:2]
allYears.sort()

print(allYears)


for year in allYears:
    xPos = 100 + (year - 1900) * 10
    d.append(draw.Text(str(year), 14, xPos, 20, fill='black', font_family='sans-serif',text_anchor='end' ))



d.set_pixel_scale(2)  # Set number of pixels per geometry unit
#d.set_render_size(400, 200)  # Alternative to set_pixel_scale
d.save_svg('example.svg')
#d.save_png('example.png')

# Display in Jupyter notebook
#d.rasterize()  # Display as PNG
# Display as SVG
